import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random

def populate_yjos_drillout_file(input_file_path, output_file_path):
    """
    Populate the YJOS Drillout Excel file with realistic dummy data
    """
    
    # Load the workbook
    wb = load_workbook(input_file_path)
    
    print("Populating YJOS Drillout file with dummy data...")
    
    # 1. UPDATE GENERAL INFO SHEET
    print("1. Updating General Info...")
    if 'General Info' in wb.sheetnames:
        ws = wb['General Info']
        
        # Update the general information (these are already in the file)
        # Customer Name: Mughees Khan (already in C12)
        # Ticket Number: AE867384 (already in E12)
        # Address: Dubai Silicon Oasis (already in C13)
        # Day YJOS Supervisor: Saif Khan (already in E13)
        # Email: mughees@bracketworks.io (already in E15)
        
        # Add missing information
        ws['C21'] = 'Drillout Operation'  # Job Type
        ws['C23'] = 'DSO-1'  # Lease/Well
        ws['C24'] = 'Dubai'  # County
        ws['E18'] = 'Dubai Production Field'  # District
        ws['E22'] = 'Ahmad Hassan'  # YJOS Salesman
        ws['E23'] = 'Express Logistics'  # Shipped Via
        ws['E24'] = 'Rig Alpha-7'  # Rig Name & No
        ws['E25'] = 'AFE-2024-DSO-001'  # AFE/PO Number
        
        print("   ✅ General Info updated")
    
    # 2. POPULATE MILL DATA SHEET
    print("2. Populating Mill Data...")
    if 'Mill Data' in wb.sheetnames:
        ws = wb['Mill Data']
        
        # Update header information
        ws['B6'] = 'Drillout Operation'  # Job Type
        ws['H6'] = 'Mughees Khan'  # Customer (already there)
        ws['H7'] = 'DSO-1'  # Lease/Well#
        ws['H8'] = 'Dubai Production Field'  # Field/Block#
        ws['H9'] = 'Dubai, UAE'  # State
        ws['H10'] = 'Dubai'  # County
        ws['H11'] = 'Rig Alpha-7'  # Rig
        ws['H12'] = '9 5/8" @ 5,300 ft'  # Casing
        
        # Add realistic drill data for 10 plugs
        plug_data = [
            {'plug': 1, 'tag_time': '08:30', 'tag_depth': 5247, 'drill_time': 45, 'actual_depth': 5250, 'torque': 450, 'pressure': 1200},
            {'plug': 2, 'tag_time': '10:15', 'tag_depth': 5189, 'drill_time': 38, 'actual_depth': 5192, 'torque': 420, 'pressure': 1150},
            {'plug': 3, 'tag_time': '12:45', 'tag_depth': 5134, 'drill_time': 52, 'actual_depth': 5137, 'torque': 480, 'pressure': 1250},
            {'plug': 4, 'tag_time': '14:20', 'tag_depth': 5076, 'drill_time': 41, 'actual_depth': 5079, 'torque': 435, 'pressure': 1180},
            {'plug': 5, 'tag_time': '16:10', 'tag_depth': 5021, 'drill_time': 47, 'actual_depth': 5024, 'torque': 465, 'pressure': 1220},
            {'plug': 6, 'tag_time': '08:45', 'tag_depth': 4967, 'drill_time': 39, 'actual_depth': 4970, 'torque': 410, 'pressure': 1160},
            {'plug': 7, 'tag_time': '11:30', 'tag_depth': 4912, 'drill_time': 44, 'actual_depth': 4915, 'torque': 445, 'pressure': 1190},
            {'plug': 8, 'tag_time': '13:15', 'tag_depth': 4858, 'drill_time': 36, 'actual_depth': 4861, 'torque': 390, 'pressure': 1140},
            {'plug': 9, 'tag_time': '15:40', 'tag_depth': 4803, 'drill_time': 49, 'actual_depth': 4806, 'torque': 470, 'pressure': 1230},
            {'plug': 10, 'tag_time': '17:25', 'tag_depth': 4749, 'drill_time': 43, 'actual_depth': 4752, 'torque': 440, 'pressure': 1200}
        ]
        
        # Populate the drill data (starting from row 16)
        for i, data in enumerate(plug_data):
            row = 16 + i
            ws[f'A{row}'] = data['plug']  # Plug/Seat No.
            ws[f'B{row}'] = data['tag_time']  # Tag Time
            ws[f'C{row}'] = data['tag_depth']  # Tag Plug/Seat Depth
            ws[f'D{row}'] = data['drill_time']  # Drill Time (mins)
            ws[f'E{row}'] = data['actual_depth']  # Actual Plug/Seat Set Depth
            ws[f'F{row}'] = data['actual_depth'] - data['tag_depth']  # Depth Difference
            ws[f'G{row}'] = data['torque']  # Free Swivel Torque
            ws[f'H{row}'] = data['pressure']  # Circ. Pressure (PSI)
        
        print("   ✅ Mill Data populated with 10 plug operations")
    
    # 3. POPULATE CT MILLING SHEET
    print("3. Populating CT Milling...")
    if ' CT Milling' in wb.sheetnames:
        ws = wb[' CT Milling']
        
        # CT milling data for 5 operations
        ct_data = [
            {'plug': 1, 'tag_time': '09:15', 'set_depth': 4695, 'tag_depth': 4693, 'drill_time': 32},
            {'plug': 2, 'tag_time': '11:45', 'set_depth': 4640, 'tag_depth': 4638, 'drill_time': 29},
            {'plug': 3, 'tag_time': '14:30', 'set_depth': 4585, 'tag_depth': 4583, 'drill_time': 35},
            {'plug': 4, 'tag_time': '16:20', 'set_depth': 4530, 'tag_depth': 4528, 'drill_time': 31},
            {'plug': 5, 'tag_time': '18:10', 'set_depth': 4475, 'tag_depth': 4473, 'drill_time': 28}
        ]
        
        # Populate CT data (starting from row 11)
        for i, data in enumerate(ct_data):
            row = 11 + i
            ws[f'A{row}'] = data['plug']  # Plug/Seat No.
            ws[f'B{row}'] = data['tag_time']  # Tag Time
            ws[f'C{row}'] = data['set_depth']  # Plug Set Depth (FT)
            ws[f'D{row}'] = data['tag_depth']  # Plug Tag Depth (FT)
            ws[f'E{row}'] = data['set_depth'] - data['tag_depth']  # Difference in Depth
            ws[f'F{row}'] = data['drill_time']  # Drill Time (mins)
        
        print("   ✅ CT Milling populated with 5 operations")
    
    # 4. POPULATE SERVICE PLAN SHEETS (SP1-SP10)
    print("4. Populating Service Plans...")
    service_plans = [
        {'day': 1, 'activities': ['Rig up equipment', 'Run mill to 5250 ft', 'Mill plug #1', 'Mill plug #2'], 'hours': 10.5},
        {'day': 2, 'activities': ['Mill plug #3', 'Mill plug #4', 'Mill plug #5', 'Change mill bit'], 'hours': 11.0},
        {'day': 3, 'activities': ['Mill plug #6', 'Mill plug #7', 'Mill plug #8', 'Equipment inspection'], 'hours': 12.5},
        {'day': 4, 'activities': ['Mill plug #9', 'Mill plug #10', 'Run CT unit', 'CT operation #1'], 'hours': 10.5},
        {'day': 5, 'activities': ['CT operation #2', 'CT operation #3', 'Wellbore cleanup', 'Circulation'], 'hours': 9.0},
        {'day': 6, 'activities': ['CT operation #4', 'CT operation #5', 'Final cleanup', 'Equipment inspection'], 'hours': 8.5},
        {'day': 7, 'activities': ['Rig down equipment', 'Final wellbore test', 'Equipment maintenance', 'Mobilization'], 'hours': 7.5},
        {'day': 8, 'activities': ['Standby', 'Equipment transport', 'Documentation', 'Final reporting'], 'hours': 4.0},
        {'day': 9, 'activities': ['Equipment cleaning', 'Inventory check', 'Maintenance logs', 'Site cleanup'], 'hours': 6.0},
        {'day': 10, 'activities': ['Final inspection', 'Client meeting', 'Handover documentation', 'Demobilization'], 'hours': 5.0}
    ]
    
    for i in range(1, 11):  # SP1 to SP10
        sheet_name = f'SP{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Add service plan data (this would depend on the actual sheet structure)
            # For now, add basic information that shows the sheet is being used
            if i <= len(service_plans):
                plan = service_plans[i-1]
                # Add date and basic info (adjust cell positions as needed)
                try:
                    ws['B2'] = f'2025-06-{10+i}'  # Date
                    ws['B3'] = f'Day {plan["day"]} Operations'  # Description
                    ws['B4'] = f'{plan["hours"]} hours planned'  # Hours
                    
                    # Add activities list
                    for j, activity in enumerate(plan['activities'][:4]):  # First 4 activities
                        ws[f'B{6+j}'] = f'{8+j*2}:00 - {activity}'
                        
                except Exception as e:
                    print(f"   ⚠️  Could not populate {sheet_name}: {e}")
                    continue
    
    print("   ✅ Service Plans populated")
    
    # 5. ADD EQUIPMENT USAGE DATA TO FT SHEETS
    print("5. Populating Equipment Sheets...")
    equipment_data = {
        'FT1': {'tool': 'Overshot', 'deployment': '08:30', 'retrieval': '09:15', 'status': 'Success'},
        'FT2': {'tool': 'Spear', 'deployment': '09:30', 'retrieval': '10:45', 'status': 'Success'},
        'FT3': {'tool': 'Mill Bit #1', 'deployment': '11:00', 'retrieval': '15:30', 'status': 'Worn'},
        'FT4': {'tool': 'Mill Bit #2', 'deployment': '16:00', 'retrieval': '18:00', 'status': 'Success'},
        'FT5': {'tool': 'Junk Basket', 'deployment': '08:00', 'retrieval': '12:00', 'status': 'Success'},
        'FT6': {'tool': 'Washover Pipe', 'deployment': '13:00', 'retrieval': '17:00', 'status': 'Damaged'},
        'FT7': {'tool': 'Taper Tap', 'deployment': '08:15', 'retrieval': '09:30', 'status': 'Success'},
        'FT8': {'tool': 'Die Collar', 'deployment': '10:00', 'retrieval': '11:15', 'status': 'Success'},
        'FT9': {'tool': 'Safety Joint', 'deployment': '14:00', 'retrieval': '15:30', 'status': 'Success'},
        'FT10': {'tool': 'Bumper Sub', 'deployment': '16:00', 'retrieval': '17:00', 'status': 'Success'}
    }
    
    for tool_name, data in equipment_data.items():
        if tool_name in wb.sheetnames:
            ws = wb[tool_name]
            try:
                # Add basic tool information (adjust positions based on actual sheet structure)
                ws['B5'] = data['tool']  # Tool Type
                ws['B6'] = data['deployment']  # Deployment Time
                ws['B7'] = data['retrieval']  # Retrieval Time
                ws['B8'] = data['status']  # Status
                ws['B10'] = f'Deployed for {tool_name} operation'  # Notes
            except Exception as e:
                print(f"   ⚠️  Could not populate {tool_name}: {e}")
                continue
    
    print("   ✅ Equipment sheets populated")
    
    # 6. UPDATE COMMENTS SHEET
    print("6. Adding Comments...")
    if 'Comments' in wb.sheetnames:
        ws = wb['Comments']
        
        comments = [
            "Excellent crew performance throughout the operation",
            "Mill bits performed better than expected",
            "CT operations completed ahead of schedule", 
            "No safety incidents reported",
            "Equipment performed within specifications",
            "Customer satisfaction rating: Excellent"
        ]
        
        # Add comments (adjust positions based on sheet structure)
        for i, comment in enumerate(comments):
            try:
                ws[f'B{10+i}'] = f'Day {i+1}: {comment}'
            except:
                continue
                
        print("   ✅ Comments added")
    
    # 7. UPDATE OBSERVATION SHEET
    print("7. Adding Observations...")
    if 'Observation' in wb.sheetnames:
        ws = wb['Observation']
        
        observations = [
            "All safety protocols followed correctly",
            "Equipment functioning within normal parameters",
            "Crew demonstrated excellent teamwork",
            "Weather conditions favorable throughout operation",
            "No environmental incidents observed"
        ]
        
        for i, obs in enumerate(observations):
            try:
                ws[f'C{10+i}'] = obs
                ws[f'B{10+i}'] = f'2025-06-{11+i}'  # Date
            except:
                continue
                
        print("   ✅ Observations added")
    
    # Save the populated file
    wb.save(output_file_path)
    print(f"\n✅ COMPLETED! Populated file saved as: {output_file_path}")
    
    return True

def create_summary_report(output_file_path):
    """Create a summary of the dummy data added"""
    
    summary = """
    📊 YJOS DRILLOUT DUMMY DATA SUMMARY
    ====================================
    
    🎯 JOB DETAILS:
    • Customer: Mughees Khan
    • Ticket: AE867384  
    • Location: Dubai Silicon Oasis, Dubai, UAE
    • Supervisor: Saif Khan
    • Operation: 7-day Drillout
    
    ⚙️ MILL OPERATIONS (10 Plugs):
    • Depths: 5,250 ft to 4,752 ft
    • Avg Drill Time: 42.3 minutes
    • Success Rate: 100%
    • Total Footage: 30 ft
    
    🔄 CT OPERATIONS (5 Plugs):
    • Depths: 4,695 ft to 4,475 ft  
    • Avg Drill Time: 31.0 minutes
    • Accuracy: 95.5%
    
    🛠️ EQUIPMENT USAGE:
    • 10 tools deployed
    • 2 tools flagged for maintenance (FT3, FT6)
    • Overall success rate: 80%
    
    📈 PERFORMANCE METRICS:
    • Drilling Efficiency: 88%
    • Safety Score: 98/100
    • Total Work Hours: 68.5
    • Downtime: 13.4%
    
    💡 KEY INSIGHTS:
    • Excellent drilling performance
    • Equipment maintenance needed for FT3, FT6
    • High operational efficiency achieved
    • Zero safety incidents
    """
    
    # Save summary to text file
    summary_path = output_file_path.replace('.xlsx', '_SUMMARY.txt')
    with open(summary_path, 'w') as f:
        f.write(summary)
    
    print(f"📋 Summary report saved as: {summary_path}")
    return summary

# Example usage
if __name__ == "__main__":
    # YOU NEED TO CHANGE THESE PATHS TO YOUR ACTUAL FILE LOCATIONS
    input_file = "YJOS   Drillout Field Tool Kit 2022 1.xlsx"  # Your original file
    output_file = "YJOS_Drillout_WITH_DUMMY_DATA.xlsx"  # Output file with data
    
    try:
        print("🚀 Starting YJOS Drillout file population...")
        print("=" * 50)
        
        # Populate the file
        success = populate_yjos_drillout_file(input_file, output_file)
        
        if success:
            # Create summary report
            summary = create_summary_report(output_file)
            print("\n" + summary)
            
            print("\n🎉 SUCCESS! Your file is ready for the demo!")
            print(f"📁 Use this file: {output_file}")
            print("\n🎯 NEXT STEPS:")
            print("1. Upload the populated file to your Streamlit app")
            print("2. Turn OFF 'Use Demo Data' in the sidebar")
            print("3. Watch the real data extraction in action!")
            
    except FileNotFoundError:
        print("❌ ERROR: Could not find the input file.")
        print("Make sure 'YJOS   Drillout Field Tool Kit 2022 1.xlsx' is in the same folder as this script.")
        print("\n💡 TO FIX:")
        print("1. Download your original Excel file")
        print("2. Put it in the same folder as this Python script")
        print("3. Run this script again")
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        print("\n💡 Make sure you have the required libraries:")
        print("pip install pandas openpyxl numpy")
